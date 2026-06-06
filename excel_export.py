"""
excel_export.py — Exports all financial data and ratios to a multi-sheet Excel file.
Returns bytes for Streamlit download button.
"""

import io
from datetime import date, datetime
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter


# Colors
DARK_BLUE  = "1B4F8A"
MID_BLUE   = "4A90D9"
LIGHT_BLUE = "D6E8FF"
GREEN      = "1E8449"
LIGHT_GREEN= "E9F7EF"
RED        = "E74C3C"
AMBER      = "F39C12"
LIGHT_GREY = "F2F4F7"
WHITE      = "FFFFFF"
BLACK      = "000000"


def _hdr_style(ws, row, col_count, bg=DARK_BLUE, fg=WHITE, bold=True):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font  = Font(bold=bold, color=fg, name="Calibri")
        cell.fill  = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_val(v, decimals=2):
    try:
        f = float(v)
        return round(f, decimals) if not np.isnan(f) else None
    except:
        return None


def _excel_safe_value(value):
    """Convert pandas/numpy/date values into openpyxl-safe scalar values."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            value = value.tz_convert(None)
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, np.generic):
        value = value.item()
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
    return value


def _excel_safe_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    safe = df.copy()
    for col in safe.columns:
        if pd.api.types.is_datetime64_any_dtype(safe[col]):
            safe[col] = pd.to_datetime(safe[col], errors="coerce").dt.tz_localize(None)
        elif safe[col].dtype == "object":
            safe[col] = safe[col].map(_excel_safe_value)
    return safe


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, freeze: bool = True,
              hdr_bg: str = DARK_BLUE, alt_bg: str = LIGHT_GREY):
    """Write a DataFrame to a worksheet with formatting."""
    df = _excel_safe_df(df)
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        cell.font      = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill      = PatternFill("solid", fgColor=hdr_bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        bg = alt_bg if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = _excel_safe_value(val)
            cell.font      = Font(name="Calibri", size=9)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                        vertical="center")
            cell.border    = _border()

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    if freeze:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def export_excel(ticker: str, data: dict, r: dict, flags: list,
                 peer_df=None, dcf_result=None, market_ready=None) -> bytes:
    """
    Generate multi-sheet Excel workbook. Returns bytes.
    Sheets: Summary | P&L | Balance Sheet | Cash Flow | Ratios | DuPont | Red Flags | Peers | DCF
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    company = data.get("company_name", ticker)
    years   = r.get("years", [])

    # ── 1. SUMMARY SHEET ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 22

    summary_items = [
        ("COMPANY OVERVIEW", None),
        ("Company", company),
        ("NSE Ticker", ticker),
        ("Sector", data.get("sector", "—")),
        ("", None),
        ("VALUATION", None),
        ("Current Price (₹)", data.get("current_price")),
        ("Market Cap (₹ Cr)", data.get("market_cap")),
        ("P/E Ratio", data.get("pe_ratio")),
        ("P/B Ratio", _fmt_val(r.get("pb_ratio"))),
        ("EV/EBITDA", _fmt_val(r.get("ev_ebitda"))),
        ("Book Value (₹)", data.get("book_value")),
        ("Dividend Yield %", data.get("dividend_yield")),
        ("52W High (₹)", data.get("high_52w")),
        ("52W Low (₹)", data.get("low_52w")),
        ("", None),
        ("GROWTH (CAGR)", None),
        ("Revenue CAGR 3Y %", _fmt_val(r.get("revenue_cagr_3y"))),
        ("Revenue CAGR 5Y %", _fmt_val(r.get("revenue_cagr_5y"))),
        ("PAT CAGR 3Y %", _fmt_val(r.get("pat_cagr_3y"))),
        ("PAT CAGR 5Y %", _fmt_val(r.get("pat_cagr_5y"))),
        ("EPS CAGR 3Y %", _fmt_val(r.get("eps_cagr_3y"))),
        ("EPS CAGR 5Y %", _fmt_val(r.get("eps_cagr_5y"))),
        ("", None),
        ("LATEST RATIOS", None),
        ("Operating Margin %", _fmt_val(r.get("operating_margin", [None])[-1] if r.get("operating_margin") else None)),
        ("Net Margin %", _fmt_val(r.get("net_margin", [None])[-1] if r.get("net_margin") else None)),
        ("ROE %", _fmt_val(r.get("roe", [None])[-1] if r.get("roe") else None)),
        ("ROCE %", _fmt_val(r.get("roce", [None])[-1] if r.get("roce") else None)),
        ("Current Ratio", _fmt_val(r.get("current_ratio", [None])[-1] if r.get("current_ratio") else None)),
        ("Debt/Equity", _fmt_val(r.get("debt_equity", [None])[-1] if r.get("debt_equity") else None)),
    ]

    if dcf_result:
        summary_items += [
            ("", None),
            ("DCF VALUATION", None),
            ("Bear Case IV (₹)", dcf_result["scenarios"]["Bear"]["intrinsic_per_share"]),
            ("Base Case IV (₹)", dcf_result["scenarios"]["Base"]["intrinsic_per_share"]),
            ("Bull Case IV (₹)", dcf_result["scenarios"]["Bull"]["intrinsic_per_share"]),
            ("WACC %", round(dcf_result["wacc_result"]["wacc"] * 100, 2)),
            ("Upside / (Downside) %", dcf_result.get("upside_pct")),
            ("Signal", dcf_result.get("signal")),
        ]

    if market_ready:
        target = market_ready.get("target", {})
        risk = market_ready.get("risk", {})
        risk_case = market_ready.get("risk_case", {})
        conf = market_ready.get("confidence", {})
        dq = market_ready.get("data_quality", {})
        tech = market_ready.get("technical", {})
        summary_items += [
            ("", None),
            ("MARKET-READY DECISION", None),
            ("Blended Target Price", target.get("target_price")),
            ("Blended Target Upside %", target.get("upside_pct")),
            ("Fundamental Rating", target.get("rating")),
            ("Target Horizon", target.get("horizon")),
            ("Risk Case Price", risk_case.get("risk_price")),
            ("Risk Case Downside %", risk_case.get("downside_pct")),
            ("Risk/Reward Ratio", risk_case.get("risk_reward_ratio")),
            ("Signal Confidence", conf.get("score")),
            ("Confidence Rating", conf.get("rating")),
            ("Risk Score", risk.get("score")),
            ("Risk Rating", risk.get("rating")),
            ("Data Quality Score", dq.get("score")),
            ("Data Quality Rating", dq.get("rating")),
            ("1Y Return %", tech.get("return_1y_pct")),
            ("1Y Volatility %", tech.get("volatility_1y_pct")),
            ("Beta vs Nifty", tech.get("beta_vs_nifty")),
        ]

    for row_idx, (label, value) in enumerate(summary_items, 1):
        a = ws_sum.cell(row=row_idx, column=1, value=label)
        b = ws_sum.cell(row=row_idx, column=2, value=value)

        if value is None and label:  # Section header
            a.font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
            a.fill  = PatternFill("solid", fgColor=DARK_BLUE)
            b.fill  = PatternFill("solid", fgColor=DARK_BLUE)
            ws_sum.merge_cells(f"A{row_idx}:B{row_idx}")
        else:
            a.font = Font(name="Calibri", size=9, bold=True)
            b.font = Font(name="Calibri", size=9)
            b.alignment = Alignment(horizontal="right")
            fill_color = LIGHT_GREY if row_idx % 2 == 0 else WHITE
            a.fill = b.fill = PatternFill("solid", fgColor=fill_color)
        a.border = b.border = _border()

    # ── 2. P&L SHEET ──
    pl_df = data.get("pl", pd.DataFrame())
    if not pl_df.empty:
        ws_pl = wb.create_sheet("P&L Statement")
        _write_df(ws_pl, pl_df)

    # ── 3. BALANCE SHEET ──
    bs_df = data.get("bs", pd.DataFrame())
    if not bs_df.empty:
        ws_bs = wb.create_sheet("Balance Sheet")
        _write_df(ws_bs, bs_df)

    # ── 4. CASH FLOW ──
    cf_df = data.get("cf", pd.DataFrame())
    if not cf_df.empty:
        ws_cf = wb.create_sheet("Cash Flow")
        _write_df(ws_cf, cf_df)

    # ── 5. RATIOS SHEET ──
    ws_rat = wb.create_sheet("All Ratios")
    ratio_rows = [
        ("── FINANCIAL DATA (₹ Cr) ──", None),
        ("Revenue",         r.get("revenue", [])),
        ("EBITDA",          r.get("ebitda", [])),
        ("PAT",             r.get("pat", [])),
        ("EPS (₹)",         r.get("eps", [])),
        ("CFO",             r.get("cfo", [])),
        ("FCF",             r.get("fcf", [])),
        ("Total Assets",    r.get("total_assets", [])),
        ("Total Debt",      r.get("total_debt", [])),
        ("Equity",          r.get("equity", [])),
        ("── PROFITABILITY (%) ──", None),
        ("Operating Margin %",  r.get("operating_margin", [])),
        ("Net Margin %",        r.get("net_margin", [])),
        ("ROE %",               r.get("roe", [])),
        ("ROA %",               r.get("roa", [])),
        ("ROCE %",              r.get("roce", [])),
        ("── LIQUIDITY ──", None),
        ("Current Ratio",   r.get("current_ratio", [])),
        ("Quick Ratio",     r.get("quick_ratio", [])),
        ("Cash Ratio",      r.get("cash_ratio", [])),
        ("── SOLVENCY ──", None),
        ("Debt/Equity",     r.get("debt_equity", [])),
        ("Interest Coverage",r.get("interest_cover", [])),
        ("Debt/EBITDA",     r.get("debt_ebitda", [])),
        ("── EFFICIENCY ──", None),
        ("Asset Turnover",  r.get("asset_turnover", [])),
        ("Receivable Days", r.get("receivable_days", [])),
        ("Payable Days",    r.get("payable_days", [])),
        ("Inventory Days",  r.get("inventory_days", [])),
        ("Cash Conversion Cycle", r.get("cash_conversion", [])),
        ("── DUPONT ──", None),
        ("DuPont Net Margin %",    r.get("dupont_net_margin", [])),
        ("DuPont Asset Turnover",  r.get("dupont_asset_turnover", [])),
        ("DuPont Equity Multiplier",r.get("dupont_equity_mult", [])),
        ("DuPont ROE %",           r.get("dupont_roe", [])),
    ]

    # Header row
    hdr_row = ["Metric"] + years
    for c_idx, h in enumerate(hdr_row, 1):
        cell = ws_rat.cell(row=1, column=c_idx, value=h)
        cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for row_idx, (label, vals) in enumerate(ratio_rows, 2):
        if vals is None:  # Section header
            cell = ws_rat.cell(row=row_idx, column=1, value=label)
            cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=9)
            cell.fill  = PatternFill("solid", fgColor=MID_BLUE)
            ws_rat.merge_cells(f"A{row_idx}:{get_column_letter(len(hdr_row))}{row_idx}")
            continue

        ws_rat.cell(row=row_idx, column=1, value=label).font = Font(bold=True, name="Calibri", size=9)
        ws_rat.cell(row=row_idx, column=1).border = _border()
        fill = PatternFill("solid", fgColor=LIGHT_GREY if row_idx % 2 == 0 else WHITE)
        ws_rat.cell(row=row_idx, column=1).fill = fill

        for c_idx, v in enumerate(vals, 2):
            cell = ws_rat.cell(row=row_idx, column=c_idx)
            cell.value  = _fmt_val(v)
            cell.font   = Font(name="Calibri", size=9)
            cell.fill   = fill
            cell.alignment = Alignment(horizontal="right")
            cell.border = _border()

    ws_rat.column_dimensions["A"].width = 30
    for i in range(2, len(years) + 2):
        ws_rat.column_dimensions[get_column_letter(i)].width = 14

    # ── 6. RED FLAGS ──
    ws_rf = wb.create_sheet("Red Flags")
    rf_headers = ["Check", "Triggered", "Severity", "Detail"]
    for c, h in enumerate(rf_headers, 1):
        cell = ws_rf.cell(row=1, column=c, value=h)
        cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        cell.border = _border()

    for r_idx, flag in enumerate(flags, 2):
        bg = "FFE0E0" if flag["triggered"] else LIGHT_GREEN
        for c_idx, val in enumerate([
            flag["check"], "YES" if flag["triggered"] else "NO",
            flag["severity"].upper(), flag["detail"]
        ], 1):
            cell = ws_rf.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = Font(name="Calibri", size=9)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True)

    ws_rf.column_dimensions["A"].width = 40
    ws_rf.column_dimensions["B"].width = 12
    ws_rf.column_dimensions["C"].width = 12
    ws_rf.column_dimensions["D"].width = 60

    # ── 7. PEERS SHEET ──
    if peer_df is not None and len(peer_df) > 0:
        ws_peer = wb.create_sheet("Peer Comparison")
        display_df = peer_df.drop(columns=["_is_target"], errors="ignore")
        _write_df(ws_peer, display_df)

    # ── 8. DCF SHEET ──
    if dcf_result:
        ws_dcf = wb.create_sheet("DCF Valuation")
        row = 1
        ws_dcf.cell(row=row, column=1, value="DCF VALUATION MODEL").font = \
            Font(bold=True, color=WHITE, name="Calibri", size=12)
        ws_dcf.cell(row=row, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws_dcf.merge_cells(f"A{row}:G{row}")
        row += 2

        for scenario in ["Bear", "Base", "Bull"]:
            s = dcf_result["scenarios"][scenario]
            fcff_df = s.get("fcff_df")
            if fcff_df is not None:
                ws_dcf.cell(row=row, column=1, value=f"{scenario} Case Projections").font = \
                    Font(bold=True, color=WHITE, name="Calibri", size=10)
                ws_dcf.cell(row=row, column=1).fill = PatternFill("solid",
                    fgColor={"Bear": RED, "Base": MID_BLUE, "Bull": GREEN}[scenario])
                ws_dcf.merge_cells(f"A{row}:G{row}")
                row += 1

                for c_idx, col in enumerate(fcff_df.columns, 1):
                    cell = ws_dcf.cell(row=row, column=c_idx, value=col)
                    cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=9)
                    cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)
                    cell.border = _border()
                row += 1

                for _, frow in fcff_df.iterrows():
                    for c_idx, val in enumerate(frow, 1):
                        cell = ws_dcf.cell(row=row, column=c_idx, value=val)
                        cell.font   = Font(name="Calibri", size=9)
                        cell.border = _border()
                    row += 1

                # Summary
                summary_dcf = [
                    ("Intrinsic Value / Share (₹)", s["intrinsic_per_share"]),
                    ("Enterprise Value (₹ Cr)", s["enterprise_value"]),
                    ("PV of Terminal Value (₹ Cr)", s["pv_terminal_value"]),
                    ("TV as % of EV", s["tv_pct_of_ev"]),
                ]
                for label, val in summary_dcf:
                    ws_dcf.cell(row=row, column=1, value=label).font = Font(bold=True, name="Calibri", size=9)
                    ws_dcf.cell(row=row, column=2, value=val)
                    row += 1
                row += 1

        ws_dcf.column_dimensions["A"].width = 28
        for i in range(2, 9):
            ws_dcf.column_dimensions[get_column_letter(i)].width = 16

    # â”€â”€ 9. MARKET-READY DECISION LAYER â”€â”€
    if market_ready:
        target = market_ready.get("target", {})
        risk = market_ready.get("risk", {})
        risk_case = market_ready.get("risk_case", {})
        conf = market_ready.get("confidence", {})
        dq = market_ready.get("data_quality", {})
        tech = market_ready.get("technical", {})
        backtest = market_ready.get("backtest", {})

        ws_dec = wb.create_sheet("Decision Thesis")
        decision_df = pd.DataFrame([
            ("Final Signal", dcf_result.get("signal") if dcf_result else None),
            ("Blended Target Price", target.get("target_price")),
            ("Blended Target Upside %", target.get("upside_pct")),
            ("Fundamental Rating", target.get("rating")),
            ("Target Horizon", target.get("horizon")),
            ("Target Methodology", target.get("methodology")),
            ("Risk Case Price", risk_case.get("risk_price")),
            ("Risk Case Downside %", risk_case.get("downside_pct")),
            ("Risk/Reward Ratio", risk_case.get("risk_reward_ratio")),
            ("Risk Haircut %", risk_case.get("haircut_pct")),
            ("Signal Confidence Score", conf.get("score")),
            ("Signal Confidence Rating", conf.get("rating")),
            ("Risk Score", risk.get("score")),
            ("Risk Rating", risk.get("rating")),
            ("Data Quality Score", dq.get("score")),
            ("Data Quality Rating", dq.get("rating")),
            ("52W Position %", tech.get("position_52w_pct")),
            ("1Y Return %", tech.get("return_1y_pct")),
            ("1Y Volatility %", tech.get("volatility_1y_pct")),
            ("Beta vs Nifty", tech.get("beta_vs_nifty")),
            ("Backtest Readiness", backtest.get("status")),
            ("Backtest Note", backtest.get("note")),
            ("Fundamental Proxy Status", backtest.get("fundamental_proxy", {}).get("status")),
            ("Fundamental Proxy Hit Rate %", backtest.get("fundamental_proxy", {}).get("hit_rate")),
        ], columns=["Metric", "Value"])
        _write_df(ws_dec, decision_df)

        ws_target = wb.create_sheet("Target Bridge")
        components = target.get("components")
        if components is not None and not components.empty:
            _write_df(ws_target, components)
        else:
            _write_df(ws_target, pd.DataFrame([{"Message": "Target bridge unavailable"}]))

        ws_risk_case = wb.create_sheet("Risk Case Price")
        risk_components = risk_case.get("components")
        if risk_components is not None and not risk_components.empty:
            _write_df(ws_risk_case, risk_components)
        else:
            _write_df(ws_risk_case, pd.DataFrame([{"Message": risk_case.get("note", "Risk case unavailable")}]))

        ws_forecast = wb.create_sheet("Forecasts")
        forecast = market_ready.get("forecast")
        if forecast is not None and not forecast.empty:
            _write_df(ws_forecast, forecast)
        else:
            _write_df(ws_forecast, pd.DataFrame([{"Message": "Forecast unavailable"}]))

        ws_triggers = wb.create_sheet("Sector Triggers")
        trigger_table = (market_ready.get("trigger_review") or {}).get("table")
        if trigger_table is not None and not trigger_table.empty:
            _write_df(ws_triggers, trigger_table)
        else:
            _write_df(ws_triggers, pd.DataFrame([{"Message": "Sector trigger review unavailable"}]))

        ws_risk = wb.create_sheet("Risk Confidence")
        risk_rows = []
        for reason in risk.get("reasons", []) or []:
            risk_rows.append({"Type": "Risk Driver", "Detail": reason})
        for warning in dq.get("warnings", []) or []:
            risk_rows.append({"Type": "Data Quality", "Detail": warning})
        for note in conf.get("notes", []) or []:
            risk_rows.append({"Type": "Confidence", "Detail": note})
        if not risk_rows:
            risk_rows.append({"Type": "Status", "Detail": "No major model-level issues"})
        _write_df(ws_risk, pd.DataFrame(risk_rows))

        ranked = market_ready.get("peer_ranking")
        if ranked is not None and not ranked.empty:
            ws_rank = wb.create_sheet("Peer Ranking")
            _write_df(ws_rank, ranked.drop(columns=["_is_target"], errors="ignore"))

        ws_src = wb.create_sheet("Data Sources")
        source_rows = []
        source_rows.append({"Field": "Report storage", "Source": "Model-versioned SQLite snapshot enabled"})
        for key, source in (data.get("sources") or {}).items():
            source_rows.append({"Field": key, "Source": source})
        for source, status in (data.get("source_status") or {}).items():
            source_rows.append({"Field": f"{source} status", "Source": status})
        for note in data.get("data_quality_notes") or []:
            source_rows.append({"Field": "Source note", "Source": note})
        if not source_rows:
            source_rows.append({"Field": "Source metadata", "Source": "Unavailable"})
        _write_df(ws_src, pd.DataFrame(source_rows))

        market_data = data.get("market_data", {}) or {}
        hist = market_data.get("price_history")
        if hist is not None and hasattr(hist, "empty") and not hist.empty:
            ws_hist = wb.create_sheet("Price History")
            _write_df(ws_hist, hist.tail(750).reset_index(drop=True))

        exchange_events = data.get("exchange_events", {}) or {}
        actions = exchange_events.get("corporate_actions")
        announcements = exchange_events.get("announcements")
        if actions is not None and hasattr(actions, "empty") and not actions.empty:
            ws_actions = wb.create_sheet("Exchange Actions")
            _write_df(ws_actions, actions.reset_index(drop=True))
        if announcements is not None and hasattr(announcements, "empty") and not announcements.empty:
            ws_ann = wb.create_sheet("Exchange Announce")
            _write_df(ws_ann, announcements.reset_index(drop=True))

        explain = market_ready.get("explainability", {})
        ws_exp = wb.create_sheet("Explainability")
        exp_rows = []
        for item in explain.get("top_reasons", []) or []:
            exp_rows.append({"Type": "Top Reason", "Detail": item})
        for item in explain.get("top_risks", []) or []:
            exp_rows.append({"Type": "Top Risk", "Detail": item})
        if not exp_rows:
            exp_rows.append({"Type": "Status", "Detail": "Explainability unavailable"})
        _write_df(ws_exp, pd.DataFrame(exp_rows))

        quarterly = market_ready.get("quarterly", {})
        qtable = quarterly.get("table")
        ws_q = wb.create_sheet("Quarterly TTM")
        if qtable is not None and not qtable.empty:
            _write_df(ws_q, qtable)
        else:
            _write_df(ws_q, pd.DataFrame([{"Message": quarterly.get("note", "Quarterly snapshot unavailable")}]))

        bands = market_ready.get("valuation_bands", {})
        btable = bands.get("table")
        ws_band = wb.create_sheet("Valuation Bands")
        if btable is not None and not btable.empty:
            _write_df(ws_band, btable)
        else:
            _write_df(ws_band, pd.DataFrame([{"Message": bands.get("note", "Valuation bands unavailable")}]))

        momentum = market_ready.get("momentum", {})
        revision = market_ready.get("earnings_revision", {})
        ws_mr = wb.create_sheet("Momentum Revision")
        mr_rows = [
            {"Area": "Momentum", "Metric": "Rating", "Value": momentum.get("rating")},
            {"Area": "Momentum", "Metric": "Score", "Value": momentum.get("score")},
            {"Area": "Revision", "Metric": "Rating", "Value": revision.get("rating")},
            {"Area": "Revision", "Metric": "Score", "Value": revision.get("score")},
        ]
        for driver in revision.get("drivers", []) or []:
            mr_rows.append({"Area": "Revision Driver", "Metric": "Driver", "Value": driver})
        _write_df(ws_mr, pd.DataFrame(mr_rows))
        mtable = momentum.get("table")
        if mtable is not None and not mtable.empty:
            ws_mtbl = wb.create_sheet("Momentum Details")
            _write_df(ws_mtbl, mtable)

        sector = market_ready.get("sector_valuation", {})
        ws_sector = wb.create_sheet("Sector Model")
        sector_rows = [
            {"Metric": "Sector Model", "Value": sector.get("sector_model")},
            {"Metric": "Target Price", "Value": sector.get("target_price")},
            {"Metric": "Upside %", "Value": sector.get("upside_pct")},
            {"Metric": "Fair Multiple", "Value": sector.get("fair_multiple")},
        ]
        for driver in sector.get("drivers", []) or []:
            sector_rows.append({"Metric": "Driver", "Value": driver})
        _write_df(ws_sector, pd.DataFrame(sector_rows))

        ws_bt = wb.create_sheet("Backtest Proxy")
        validation = backtest.get("historical_validation", {}) or backtest.get("price_proxy", {})
        summary = validation.get("summary")
        if summary is not None and not summary.empty:
            _write_df(ws_bt, summary)
        else:
            _write_df(ws_bt, pd.DataFrame([{"Message": validation.get("note", "Historical validation unavailable")}]))
        trades = validation.get("trades")
        if trades is not None and not trades.empty:
            ws_trades = wb.create_sheet("Validation Trades")
            _write_df(ws_trades, trades)
        by_regime = validation.get("by_regime")
        if by_regime is not None and not by_regime.empty:
            ws_regime = wb.create_sheet("Validation Regime")
            _write_df(ws_regime, by_regime)
        by_sector = validation.get("by_sector")
        if by_sector is not None and not by_sector.empty:
            ws_vsector = wb.create_sheet("Validation Sector")
            _write_df(ws_vsector, by_sector)
        by_mcap = validation.get("by_market_cap")
        if by_mcap is not None and not by_mcap.empty:
            ws_vmcap = wb.create_sheet("Validation MCap")
            _write_df(ws_vmcap, by_mcap)

        ws_method = wb.create_sheet("Methodology")
        method_rows = [
            ("Fundamentals", "Screener.in statements are parsed and ratios are recalculated internally."),
            ("Market Data", "Yahoo Finance supplies price history, returns, volatility and beta when available."),
            ("DCF", "FCFF is projected using derived growth, margin, tax, capex and working-capital assumptions."),
            ("Monte Carlo", "Valuation is simulated across growth, WACC, terminal growth, margin and reinvestment uncertainty."),
            ("Target Price", "Blends DCF, fair P/E, fair P/B, peer multiples and sector/platform-specific models where applicable."),
            ("Quarterly Results", "Latest quarterly revenue, PAT, EPS, TTM values, YoY/QoQ growth and margin trend are integrated when available."),
            ("Valuation Bands", "Historical price-derived P/E/P/B proxy bands and valuation percentiles are used to flag valuation stretch."),
            ("Earnings Revision", "Growth acceleration, latest quarterly growth and margin change form an upgrade/downgrade proxy."),
            ("Momentum", "50 DMA, 200 DMA, relative strength, drawdown and volume trends are used as a secondary timing overlay."),
            ("Historical Validation", "Rolling historical signal states are tested for 1M/3M/6M/12M returns, hit rate, excess return, drawdown, regime and market-cap breakdowns."),
            ("Risk Score", "Combines red flags, leverage, volatility, valuation stretch, business type and data completeness."),
            ("Confidence", "Combines data quality, composite signal strength, Monte Carlo agreement and risk level."),
            ("Limitations", "Automated research aid only; not investment advice or audited performance evidence. Consult a SEBI-registered investment advisor or qualified financial professional before making investment decisions."),
        ]
        _write_df(ws_method, pd.DataFrame(method_rows, columns=["Area", "Methodology"]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
